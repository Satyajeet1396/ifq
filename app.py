import streamlit as st
import pandas as pd
from rapidfuzz import process, fuzz
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from io import BytesIO
from functools import lru_cache
from tqdm import tqdm
import qrcode
import base64
import openpyxl

# Reuse the existing journal standardization functions
JOURNAL_REPLACEMENTS = {
    'intl': 'international',
    'int': 'international',
    'natl': 'national',
    'nat': 'national',
    'sci': 'science',
    'med': 'medical',
    'res': 'research',
    'tech': 'technology',
    'eng': 'engineering',
    'phys': 'physics',
    'chem': 'chemistry',
    'bio': 'biology',
    'env': 'environmental',
    'mgmt': 'management',
    'dev': 'development',
    'edu': 'education',
    'univ': 'university',
    'j\\.': 'journal',
    'jr\\.': 'journal',
    'jrnl\\.': 'journal',
    'proc\\.': 'proceedings',
    'rev\\.': 'review',
    'q\\.': 'quarterly',
}

@lru_cache(maxsize=10000)
def standardize_text(text):
    if not isinstance(text, str):
        return ""
    text = text.lower().strip()
    text = text.replace('&', 'and')
    text = re.sub(r'[-:]', ' ', text)
    text = re.sub(r'\([^)]*?(print|online|www|web|issn|doi).*?\)', '', text, flags=re.IGNORECASE)
    for abbr, full in JOURNAL_REPLACEMENTS.items():
        text = re.sub(rf'\b{abbr}\b', full, text, flags=re.IGNORECASE)
    text = re.sub(r'[^\w\s]', ' ', text)
    return ' '.join(text.split())

def read_file(file):
    file_extension = file.name.split('.')[-1].lower()
    if file_extension == 'xlsx':
        return pd.read_excel(file)
    elif file_extension == 'csv':
        return pd.read_csv(file)
    else:
        raise ValueError("Unsupported file format. Please upload either CSV or XLSX file.")

def process_single_file(user_df, ref_df):
    # Find the "Source title" column
    source_title_col = None
    for col in user_df.columns:
        if 'source title' in str(col).lower():
            source_title_col = col
            break
    
    if source_title_col is None:
        st.error("No 'Source title' column found in the input file. Please ensure your file has a column containing 'Source title'.")
        return None
    
    # Create a copy of journal names for processing
    journals = user_df[source_title_col].astype(str).apply(standardize_text)
    
    # Create reference dictionary with both SJR Best Quartile and JIF values
    ref_dict = {}
    ref_df.iloc[:, 0] = ref_df.iloc[:, 0].astype(str).apply(standardize_text)  # Standardize reference journal names
    
    # Assuming ref_df has columns: [Journal Name, SJR Best Quartile, JIF]
    # If not, adjust the column indices accordingly
    for i, row in ref_df.iterrows():
        journal = row.iloc[0]  # Journal name
        sjr = row.iloc[1] if len(row) > 1 else "N/A"  # SJR Best Quartile
        jif = row.iloc[2] if len(row) > 2 else "N/A"  # JIF
        
        if journal not in ref_dict:
            ref_dict[journal] = {"sjr": [], "jif": []}
        
        ref_dict[journal]["sjr"].append(sjr)
        ref_dict[journal]["jif"].append(jif)
    
    ref_journals = ref_df.iloc[:, 0].tolist()
    journal_list = journals.tolist()
    
    results = []
    for journal in tqdm(journal_list, desc="Processing journals"):
        if pd.isna(journal) or str(journal).strip() == "":
            results.append((journal, "No match found", 0, "N/A", "N/A"))
            continue
            
        if journal in ref_dict:
            results.append((
                journal, 
                journal, 
                100, 
                ', '.join(map(str, ref_dict[journal]["sjr"])), 
                ', '.join(map(str, ref_dict[journal]["jif"]))
            ))
            continue
            
        match = process.extractOne(journal, ref_journals, scorer=fuzz.ratio, score_cutoff=90)
        if match:
            matched_journal = match[0]
            results.append((
                journal, 
                matched_journal, 
                match[1], 
                ', '.join(map(str, ref_dict[matched_journal]["sjr"])), 
                ', '.join(map(str, ref_dict[matched_journal]["jif"]))
            ))
        else:
            results.append((journal, "No match found", 0, "N/A", "N/A"))
    
    # Create DataFrame with match results including SJR and JIF
    new_columns = ['Processed Journal Name', 'Best Match', 'Match Score', 'SJR Best Quartile', 'JIF']
    results_df = pd.DataFrame(results, columns=new_columns)
    
    # Print matching statistics
    total = len(results_df)
    perfect_matches = len(results_df[results_df['Match Score'] == 100])
    good_matches = len(results_df[(results_df['Match Score'] >= 90) & (results_df['Match Score'] < 100)])
    no_matches = len(results_df[results_df['Match Score'] == 0])
    
    st.write("### Matching Statistics")
    st.write(f"""
    - Total journals: {total}
    - Perfect matches (100): {perfect_matches} ({perfect_matches/total*100:.1f}%)
    - Good matches (90-99): {good_matches} ({good_matches/total*100:.1f}%)
    - No matches: {no_matches} ({no_matches/total*100:.1f}%)
    """)
    
    # Add processed journal name and match results
    final_df = pd.concat([
        user_df,
        results_df
    ], axis=1)
    
    # Sort by Match Score in ascending order
    final_df = final_df.sort_values(by='Match Score', ascending=True)
    
    # Store the new column names for highlighting in Excel
    final_df.attrs['new_columns'] = new_columns
    
    return final_df

def save_results(df, file_format='xlsx'):
    output = BytesIO()
    
    if file_format == 'xlsx':
        # Save to Excel with styled headers
        writer = pd.ExcelWriter(output, engine='openpyxl')
        df.to_excel(writer, index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        
        # Create fill styles for different columns
        header_fill = PatternFill(start_color='0066CC',
                                end_color='0066CC',
                                fill_type='solid')
        
        sjr_fill = PatternFill(start_color='00CC66',
                             end_color='00CC66',
                             fill_type='solid')
                             
        jif_fill = PatternFill(start_color='FF9900',
                             end_color='FF9900',
                             fill_type='solid')
        
        # Get new column names from DataFrame attributes
        new_columns = df.attrs.get('new_columns', [])
        
        # Apply highlighting to column headers
        for cell in worksheet[1]:
            if cell.value in new_columns:
                if cell.value == 'SJR Best Quartile':
                    cell.fill = sjr_fill
                    cell.font = Font(color='000000', bold=True)
                elif cell.value == 'JIF':
                    cell.fill = jif_fill
                    cell.font = Font(color='000000', bold=True)
                else:
                    cell.fill = header_fill
                    cell.font = Font(color='FFFFFF', bold=True)
        
        # Save the workbook
        writer.close()
    else:
        # For CSV, just save normally
        df.to_csv(output, index=False)
    
    output.seek(0)
    return output

# Streamlit app
st.title("Journal Impact Factor and SJR Quartile Processor")

# Add collapsible app information
with st.expander("ℹ️ Click here to learn about this app", expanded=False):
    st.markdown("""
        <style>
        .app-info {
            padding: 20px;
            border-radius: 10px;
            background-color: #f0f2f6;
            margin: 10px 0;
        }
        .app-info h3 {
            color: #0066cc;
            margin-top: 20px;
            margin-bottom: 10px;
        }
        .app-info ul, .app-info ol {
            margin-bottom: 20px;
        }
        </style>
        <div class="app-info">
        <h3>📚 About This App</h3>
        <p>This app helps you find impact factors and SJR quartiles for your journal lists. It can:</p>
        <ul>
        <li>Process multiple Excel/CSV files at once</li>
        <li>Automatically finds the 'Source title' column in your data</li>
        <li>Handle journal name variations and abbreviations</li>
        <li>Display both SJR Best Quartile and JIF values</li>
        <li>Sort results by match quality (poorest matches first)</li>
        <li>Preserves all original columns and adds match results at the end</li>
        </ul>
        
        <h3>🔍 How to Use</h3>
        <ol>
        <li>Upload one or more Excel/CSV files containing a 'Source title' column</li>
        <li>Wait for processing to complete</li>
        <li>Review results (sorted with poorest matches first)</li>
        <li>Download processed results for each file</li>
        </ol>
        
        <h3>📊 Match Score Guide</h3>
        <ul>
        <li><strong>100</strong>: Perfect match</li>
        <li><strong>90-99</strong>: Good match with minor variations</li>
        <li><strong>0</strong>: No match found</li>
        </ul>
        </div>
    """, unsafe_allow_html=True)

st.write("Upload multiple journal lists to process them simultaneously.")

# Initialize session states
if 'processed_files' not in st.session_state:
    st.session_state.processed_files = set()
if 'processed_results' not in st.session_state:
    st.session_state.processed_results = {}

# File uploads
uploaded_files = st.file_uploader("Upload Your Journal Lists (Excel/CSV)", type=["xlsx", "csv"], accept_multiple_files=True, key="file_uploader")

# Reference file loading with error handling
try:
    reference_file_url = "https://github.com/Satyajeet1396/ifq/blob/90057686057c21c7aaff527c1a58dfc644587fbd/ifq.xlsx"
    ref_df = pd.read_excel(reference_file_url)
    st.write(f"Successfully loaded reference database with {len(ref_df)} entries")
except Exception as e:
    st.error(f"Error loading reference database: {str(e)}")
    st.stop()

if uploaded_files:
    # Process each file that hasn't been processed yet
    for uploaded_file in uploaded_files:
        file_identifier = f"{uploaded_file.name}_{uploaded_file.size}"
        
        if file_identifier not in st.session_state.processed_files:
            st.write(f"Processing {uploaded_file.name}...")
            
            try:
                # Read and process the file
                user_df = read_file(uploaded_file)
                st.write(f"Found {len(user_df)} entries in {uploaded_file.name}")
                
                with st.spinner(f"Processing {uploaded_file.name}..."):
                    results_df = process_single_file(user_df, ref_df)
                    
                    # Save results to session state
                    output_format = uploaded_file.name.split('.')[-1].lower()
                    output_file = save_results(results_df, output_format)
                    
                    st.session_state.processed_results[file_identifier] = {
                        'results_df': results_df,
                        'output_file': output_file,
                        'output_format': output_format,
                        'filename': uploaded_file.name
                    }
                    
                    # Mark file as processed
                    st.session_state.processed_files.add(file_identifier)
            
            except Exception as e:
                st.error(f"Error processing {uploaded_file.name}: {str(e)}")
                continue
    
    # Display results for all processed files
    if st.session_state.processed_results:
        st.write("### Processed Files Results")
        for file_id, data in st.session_state.processed_results.items():
            with st.expander(f"Results for {data['filename']}", expanded=True):
                # Create download button for this file
                output_filename = f"{data['filename'].rsplit('.', 1)[0]}_matched.{data['output_format']}"
                mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if data['output_format'] == 'xlsx' else "text/csv"
                
                st.download_button(
                    label=f"Download Results for {data['filename']}",
                    data=data['output_file'],
                    file_name=output_filename,
                    mime=mime_type,
                    key=f"download_{file_id}"
                )
                
                # Show sample results with focus on SJR and JIF columns
                st.write(f"Sample results:")
                
                # Display a subset of columns with SJR and JIF highlighted
                display_cols = ['Processed Journal Name', 'Best Match', 'Match Score', 'SJR Best Quartile', 'JIF']
                
                # Check if these columns exist in the results
                available_cols = [col for col in display_cols if col in data['results_df'].columns]
                
                # Display the dataframe with the available columns
                st.dataframe(data['results_df'][available_cols].head())
                
                # Add a specific section to highlight SJR and JIF values
                st.write("### SJR Best Quartile and JIF Values")
                sjr_jif_df = data['results_df'][['Best Match', 'SJR Best Quartile', 'JIF']].head(10)
                st.dataframe(sjr_jif_df, use_container_width=True)
    
    # Add a button to clear processed files and start fresh
    if st.button("Clear All and Process New Files"):
        st.session_state.processed_files.clear()
        st.session_state.processed_results.clear()
        st.experimental_rerun()
            
else:
    st.info("Please upload one or more journal lists (XLSX or CSV format) to get started.")

st.divider()
st.info("Created by Dr. Satyajeet Patil")
st.info("For more cool apps like this visit: https://patilsatyajeet.wixsite.com/home/python")

# Support section in expander
with st.expander("🤝 Support Our Research", expanded=False):
    st.markdown("""
        <div style='text-align: center; padding: 1rem; background-color: #f0f2f6; border-radius: 10px; margin: 1rem 0;'>
            <h3>🙏 Your Support Makes a Difference!</h3>
            <p>Your contribution helps us continue developing free tools for the research community.</p>
            <p>Every donation, no matter how small, fuels our research journey!</p>
        </div>
        """, unsafe_allow_html=True)
    
    # Two columns for QR code and Buy Me a Coffee button
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("#### UPI Payment")
        # Generate UPI QR code
        upi_url = "upi://pay?pa=satyajeet1396@oksbi&pn=Satyajeet Patil&cu=INR"
        qr = qrcode.make(upi_url)
        
        # Save QR code to BytesIO
        buffer = BytesIO()
        qr.save(buffer, format="PNG")
        buffer.seek(0)
        qr_base64 = base64.b64encode(buffer.getvalue()).decode()
        
        # Display QR code with message
        st.markdown("Scan to pay: **satyajeet1396@oksbi**")
        st.markdown(
            f"""
            <div style="display: flex; justify-content: center; align-items: center;">
                <img src="data:image/png;base64,{qr_base64}" width="200">
            </div>
            """,
            unsafe_allow_html=True
        )
    
    with col2:
        st.markdown("#### Buy Me a Coffee")
        st.markdown("Support through Buy Me a Coffee platform:")
        # Buy Me a Coffee button
        st.markdown(
            """
            <div style="display: flex; justify-content: center; align-items: center; height: 100%;">
                <a href="https://www.buymeacoffee.com/researcher13" target="_blank">
                    <img src="https://img.buymeacoffee.com/button-api/?text=Support our Research&emoji=&slug=researcher13&button_colour=FFDD00&font_colour=000000&font_family=Cookie&outline_colour=000000&coffee_colour=ffffff" alt="Support our Research"/>
                </a>
            </div>
            """,
            unsafe_allow_html=True
        )

st.info("A small donation from you can fuel our research journey, turning ideas into breakthroughs that can change lives!")
